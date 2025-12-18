"""
Generate comprehensive tower location reports in multiple formats:
CSV, JSON, GeoJSON, Excel, and PDF
"""

import pandas as pd
import json
import logging
from pathlib import Path
from typing import Dict, List, Optional
from datetime import datetime
import numpy as np
import sys

# Optional imports for advanced formats
try:
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment
    from openpyxl.utils.dataframe import dataframe_to_rows
    OPENPYXL_AVAILABLE = True
except ImportError:
    OPENPYXL_AVAILABLE = False
    logging.warning("openpyxl not available - Excel export will be skipped")

try:
    from reportlab.lib.pagesizes import letter, A4
    from reportlab.lib import colors
    from reportlab.lib.units import inch
    from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer, PageBreak
    from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
    REPORTLAB_AVAILABLE = True
except ImportError:
    REPORTLAB_AVAILABLE = False
    logging.warning("reportlab not available - PDF export will be skipped")

# Setup logging
logging.basicConfig(
    level=logging.INFO,
    format='%(asctime)s - %(name)s - %(levelname)s - %(message)s'
)
logger = logging.getLogger(__name__)

PROJECT_ROOT = Path(__file__).parent.parent
DATA_DIR = PROJECT_ROOT / "data"
OUTPUT_DIR = DATA_DIR / "outputs" / "tower_locations"
OUTPUT_DIR.mkdir(parents=True, exist_ok=True)


class TowerLocationReportGenerator:
    """Generate comprehensive reports in multiple formats"""
    
    def __init__(self, df: pd.DataFrame):
        self.df = df.copy()
        self.output_dir = OUTPUT_DIR
        self.timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
        
    def generate_csv_report(self) -> Path:
        """Generate CSV report"""
        logger.info("Generating CSV report...")
        
        filename = f"complete_inventory_{self.timestamp}.csv"
        filepath = self.output_dir / filename
        
        self.df.to_csv(filepath, index=False, encoding='utf-8')
        
        logger.info(f"✓ CSV report saved: {filepath}")
        return filepath
    
    def generate_json_report(self) -> Path:
        """Generate JSON report"""
        logger.info("Generating JSON report...")
        
        filename = f"complete_inventory_{self.timestamp}.json"
        filepath = self.output_dir / filename
        
        # Convert DataFrame to records
        records = self.df.to_dict('records')
        
        # Add metadata
        report_data = {
            'metadata': {
                'generation_date': datetime.now().isoformat(),
                'total_towers': len(self.df),
                'regions': self.df['region'].nunique() if 'region' in self.df.columns else 0,
                'states': self.df['state_code'].nunique() if 'state_code' in self.df.columns else 0,
            },
            'towers': records
        }
        
        with open(filepath, 'w', encoding='utf-8') as f:
            json.dump(report_data, f, indent=2, ensure_ascii=False, default=str)
        
        logger.info(f"✓ JSON report saved: {filepath}")
        return filepath
    
    def generate_geojson_report(self) -> Path:
        """Generate GeoJSON report for mapping/GIS"""
        logger.info("Generating GeoJSON report...")
        
        filename = f"complete_inventory_{self.timestamp}.geojson"
        filepath = self.output_dir / filename
        
        features = []
        
        for _, row in self.df.iterrows():
            feature = {
                'type': 'Feature',
                'properties': {
                    'tower_id': str(row.get('tower_id', '')),
                    'maintenance_zone': str(row.get('maintenance_zone', '')),
                    'zone_type': str(row.get('zone_type', '')),
                    'region': str(row.get('region', '')),
                    'state_code': str(row.get('state_code', '')),
                    'state_name': str(row.get('state_name', '')),
                    'tower_type': str(row.get('tower_type', '')),
                    'status': str(row.get('status', '')),
                    'priority': str(row.get('priority', '')),
                    'height_meters': float(row.get('height_meters', 0)) if pd.notna(row.get('height_meters')) else None,
                    'operator_count': int(row.get('operator_count', 0)) if pd.notna(row.get('operator_count')) else None,
                    'signal_strength': int(row.get('signal_strength', 0)) if pd.notna(row.get('signal_strength')) else None,
                    'uptime_percent': float(row.get('uptime_percent', 0)) if pd.notna(row.get('uptime_percent')) else None,
                },
                'geometry': {
                    'type': 'Point',
                    'coordinates': [float(row['longitude']), float(row['latitude'])]
                }
            }
            features.append(feature)
        
        geojson_data = {
            'type': 'FeatureCollection',
            'metadata': {
                'generation_date': datetime.now().isoformat(),
                'total_towers': len(self.df),
            },
            'features': features
        }
        
        with open(filepath, 'w', encoding='utf-8') as f:
            json.dump(geojson_data, f, indent=2, ensure_ascii=False)
        
        logger.info(f"✓ GeoJSON report saved: {filepath}")
        return filepath
    
    def generate_excel_report(self) -> Optional[Path]:
        """Generate Excel report with multiple sheets"""
        if not OPENPYXL_AVAILABLE:
            logger.warning("openpyxl not available - skipping Excel report")
            return None
        
        logger.info("Generating Excel report...")
        
        filename = f"complete_inventory_{self.timestamp}.xlsx"
        filepath = self.output_dir / filename
        
        wb = Workbook()
        
        # Remove default sheet
        if 'Sheet' in wb.sheetnames:
            wb.remove(wb['Sheet'])
        
        # Sheet 1: Complete Inventory
        ws1 = wb.create_sheet("Complete Inventory", 0)
        self._write_dataframe_to_sheet(ws1, self.df, "Complete Tower Inventory")
        
        # Sheet 2: Summary by Region
        if 'region' in self.df.columns:
            ws2 = wb.create_sheet("Summary by Region", 1)
            region_summary = self._create_region_summary()
            self._write_dataframe_to_sheet(ws2, region_summary, "Tower Distribution by Region")
        
        # Sheet 3: Summary by State
        if 'state_code' in self.df.columns:
            ws3 = wb.create_sheet("Summary by State", 2)
            state_summary = self._create_state_summary()
            self._write_dataframe_to_sheet(ws3, state_summary, "Tower Distribution by State")
        
        # Sheet 4: Summary by Maintenance Zone
        if 'maintenance_zone' in self.df.columns:
            ws4 = wb.create_sheet("Summary by Zone", 3)
            zone_summary = self._create_zone_summary()
            self._write_dataframe_to_sheet(ws4, zone_summary, "Tower Distribution by Maintenance Zone")
        
        # Sheet 5: Statistics
        ws5 = wb.create_sheet("Statistics", 4)
        self._write_statistics_to_sheet(ws5)
        
        wb.save(filepath)
        
        logger.info(f"✓ Excel report saved: {filepath}")
        return filepath
    
    def _write_dataframe_to_sheet(self, ws, df: pd.DataFrame, title: str):
        """Write DataFrame to Excel sheet with formatting"""
        # Write title
        ws['A1'] = title
        ws['A1'].font = Font(bold=True, size=14)
        ws.merge_cells('A1:Z1')
        
        # Write headers
        headers = list(df.columns)
        for col_idx, header in enumerate(headers, start=1):
            cell = ws.cell(row=2, column=col_idx, value=header)
            cell.font = Font(bold=True)
            cell.fill = PatternFill(start_color="CCCCCC", end_color="CCCCCC", fill_type="solid")
            cell.alignment = Alignment(horizontal="center")
        
        # Write data
        for row_idx, row_data in enumerate(dataframe_to_rows(df, index=False, header=False), start=3):
            for col_idx, value in enumerate(row_data, start=1):
                # Convert dict/list to string for Excel compatibility
                if isinstance(value, (dict, list)):
                    value = json.dumps(value, ensure_ascii=False) if value else ''
                ws.cell(row=row_idx, column=col_idx, value=value)
        
        # Auto-adjust column widths
        from openpyxl.utils import get_column_letter
        for col_idx, column in enumerate(ws.columns, start=1):
            max_length = 0
            for cell in column:
                try:
                    if cell.value and len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            column_letter = get_column_letter(col_idx)
            adjusted_width = min(max_length + 2, 50)
            ws.column_dimensions[column_letter].width = adjusted_width
    
    def _create_region_summary(self) -> pd.DataFrame:
        """Create summary statistics by region"""
        if 'region' not in self.df.columns:
            return pd.DataFrame()
        
        summary = self.df.groupby('region').agg({
            'tower_id': 'count',
            'status': lambda x: json.dumps(x.value_counts().to_dict(), ensure_ascii=False),
            'priority': lambda x: json.dumps(x.value_counts().to_dict(), ensure_ascii=False),
        }).reset_index()
        
        summary.columns = ['region', 'total_towers', 'status_distribution', 'priority_distribution']
        
        return summary
    
    def _create_state_summary(self) -> pd.DataFrame:
        """Create summary statistics by state"""
        if 'state_code' not in self.df.columns:
            return pd.DataFrame()
        
        summary = self.df.groupby(['state_code', 'state_name']).agg({
            'tower_id': 'count',
            'maintenance_zone': 'nunique',
        }).reset_index()
        
        summary.columns = ['state_code', 'state_name', 'total_towers', 'zones_count']
        summary = summary.sort_values('total_towers', ascending=False)
        
        return summary
    
    def _create_zone_summary(self) -> pd.DataFrame:
        """Create summary statistics by maintenance zone"""
        if 'maintenance_zone' not in self.df.columns:
            return pd.DataFrame()
        
        summary = self.df.groupby(['maintenance_zone', 'region', 'zone_type']).agg({
            'tower_id': 'count',
            'status': lambda x: json.dumps(x.value_counts().to_dict(), ensure_ascii=False),
        }).reset_index()
        
        summary.columns = ['maintenance_zone', 'region', 'zone_type', 'total_towers', 'status_distribution']
        summary = summary.sort_values('total_towers', ascending=False)
        
        return summary
    
    def _write_statistics_to_sheet(self, ws):
        """Write overall statistics to Excel sheet"""
        ws['A1'] = "Nova Corrente Tower Location Report - Statistics"
        ws['A1'].font = Font(bold=True, size=14)
        ws.merge_cells('A1:B1')
        
        row = 3
        stats = self._calculate_statistics()
        
        for key, value in stats.items():
            ws.cell(row=row, column=1, value=str(key)).font = Font(bold=True)
            ws.cell(row=row, column=2, value=str(value))
            row += 1
    
    def _calculate_statistics(self) -> Dict:
        """Calculate overall statistics"""
        stats = {
            'Total Towers': len(self.df),
            'Generation Date': datetime.now().strftime('%Y-%m-%d %H:%M:%S'),
        }
        
        if 'region' in self.df.columns:
            stats['Regions Covered'] = self.df['region'].nunique()
            stats['Towers by Region'] = self.df['region'].value_counts().to_dict()
        
        if 'state_code' in self.df.columns:
            stats['States Covered'] = self.df['state_code'].nunique()
        
        if 'maintenance_zone' in self.df.columns:
            stats['Maintenance Zones'] = self.df['maintenance_zone'].nunique()
        
        if 'status' in self.df.columns:
            stats['Status Distribution'] = self.df['status'].value_counts().to_dict()
        
        if 'priority' in self.df.columns:
            stats['Priority Distribution'] = self.df['priority'].value_counts().to_dict()
        
        return stats
    
    def generate_pdf_report(self) -> Optional[Path]:
        """Generate PDF report"""
        if not REPORTLAB_AVAILABLE:
            logger.warning("reportlab not available - skipping PDF report")
            return None
        
        logger.info("Generating PDF report...")
        
        filename = f"complete_inventory_{self.timestamp}.pdf"
        filepath = self.output_dir / filename
        
        doc = SimpleDocTemplate(str(filepath), pagesize=A4)
        story = []
        styles = getSampleStyleSheet()
        
        # Title
        title_style = ParagraphStyle(
            'CustomTitle',
            parent=styles['Heading1'],
            fontSize=18,
            textColor=colors.HexColor('#1a1a1a'),
            spaceAfter=30,
            alignment=1  # Center
        )
        story.append(Paragraph("Nova Corrente Tower Location Report", title_style))
        story.append(Spacer(1, 0.2*inch))
        
        # Executive Summary
        story.append(Paragraph("Executive Summary", styles['Heading2']))
        summary_text = f"""
        This report contains the complete inventory of {len(self.df):,} telecommunications towers 
        under Nova Corrente maintenance across Brazil. The towers are distributed across 
        {self.df['region'].nunique() if 'region' in self.df.columns else 'N/A'} regions and 
        {self.df['state_code'].nunique() if 'state_code' in self.df.columns else 'N/A'} states.
        """
        story.append(Paragraph(summary_text, styles['Normal']))
        story.append(Spacer(1, 0.2*inch))
        
        # Statistics Table
        story.append(Paragraph("Key Statistics", styles['Heading2']))
        stats = self._calculate_statistics()
        
        stats_data = [['Metric', 'Value']]
        for key, value in list(stats.items())[:10]:  # First 10 stats
            if isinstance(value, dict):
                value = json.dumps(value, ensure_ascii=False)[:50] + "..."
            stats_data.append([str(key), str(value)])
        
        stats_table = Table(stats_data)
        stats_table.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), colors.grey),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
            ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (-1, 0), 12),
            ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
            ('BACKGROUND', (0, 1), (-1, -1), colors.beige),
            ('GRID', (0, 0), (-1, -1), 1, colors.black)
        ]))
        story.append(stats_table)
        story.append(Spacer(1, 0.3*inch))
        
        # Regional Distribution
        if 'region' in self.df.columns:
            story.append(Paragraph("Tower Distribution by Region", styles['Heading2']))
            region_data = [['Region', 'Tower Count', 'Percentage']]
            
            region_counts = self.df['region'].value_counts()
            total = len(self.df)
            
            for region, count in region_counts.items():
                pct = (count / total) * 100
                region_data.append([str(region), str(count), f"{pct:.1f}%"])
            
            region_table = Table(region_data)
            region_table.setStyle(TableStyle([
                ('BACKGROUND', (0, 0), (-1, 0), colors.grey),
                ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
                ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
                ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
                ('FONTSIZE', (0, 0), (-1, 0), 12),
                ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
                ('BACKGROUND', (0, 1), (-1, -1), colors.beige),
                ('GRID', (0, 0), (-1, -1), 1, colors.black)
            ]))
            story.append(region_table)
            story.append(PageBreak())
        
        # State Distribution
        if 'state_code' in self.df.columns:
            story.append(Paragraph("Tower Distribution by State", styles['Heading2']))
            state_data = [['State Code', 'State Name', 'Tower Count']]
            
            state_counts = self.df.groupby(['state_code', 'state_name']).size().reset_index(name='count')
            state_counts = state_counts.sort_values('count', ascending=False)
            
            for _, row in state_counts.head(15).iterrows():  # Top 15 states
                state_data.append([
                    str(row['state_code']),
                    str(row['state_name']),
                    str(row['count'])
                ])
            
            state_table = Table(state_data)
            state_table.setStyle(TableStyle([
                ('BACKGROUND', (0, 0), (-1, 0), colors.grey),
                ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
                ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
                ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
                ('FONTSIZE', (0, 0), (-1, 0), 10),
                ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
                ('BACKGROUND', (0, 1), (-1, -1), colors.beige),
                ('GRID', (0, 0), (-1, -1), 1, colors.black)
            ]))
            story.append(state_table)
        
        # Footer
        story.append(Spacer(1, 0.5*inch))
        footer_text = f"Report generated on {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}"
        story.append(Paragraph(footer_text, styles['Normal']))
        
        doc.build(story)
        
        logger.info(f"✓ PDF report saved: {filepath}")
        return filepath
    
    def generate_all_reports(self) -> Dict[str, Path]:
        """Generate all report formats"""
        logger.info("=" * 80)
        logger.info("GENERATING ALL REPORT FORMATS")
        logger.info("=" * 80)
        
        reports = {}
        
        # Generate CSV
        reports['csv'] = self.generate_csv_report()
        
        # Generate JSON
        reports['json'] = self.generate_json_report()
        
        # Generate GeoJSON
        reports['geojson'] = self.generate_geojson_report()
        
        # Generate Excel
        excel_path = self.generate_excel_report()
        if excel_path:
            reports['excel'] = excel_path
        
        # Generate PDF
        pdf_path = self.generate_pdf_report()
        if pdf_path:
            reports['pdf'] = pdf_path
        
        logger.info("=" * 80)
        logger.info("ALL REPORTS GENERATED SUCCESSFULLY")
        logger.info("=" * 80)
        
        for format_type, filepath in reports.items():
            logger.info(f"  {format_type.upper()}: {filepath}")
        
        return reports


def main():
    """Main execution function"""
    # Load consolidated tower data
    tower_files = list(OUTPUT_DIR.glob("complete_tower_inventory_*.csv"))
    
    if not tower_files:
        logger.error("No tower inventory file found. Run extract_tower_locations.py first.")
        return
    
    # Use most recent file
    latest_file = max(tower_files, key=lambda p: p.stat().st_mtime)
    logger.info(f"Loading tower data from: {latest_file}")
    
    df = pd.read_csv(latest_file)
    logger.info(f"Loaded {len(df)} towers")
    
    # Generate all reports
    generator = TowerLocationReportGenerator(df)
    reports = generator.generate_all_reports()
    
    logger.info(f"\n✓ Generated {len(reports)} report formats")
    return reports


if __name__ == '__main__':
    main()

